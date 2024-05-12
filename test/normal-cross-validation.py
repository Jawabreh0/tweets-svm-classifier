import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import SVC
import joblib

# Load the trained model and vectorizer
model = joblib.load('./trained_models/classifier.pkl')
vectorizer = joblib.load('./trained_models/vectorizer.pkl')

def predict_tweet(tweet):
    # Clean and prepare the tweet
    tweet_cleaned = tweet.strip()
    tweet_vectorized = vectorizer.transform([tweet_cleaned])
    return model.predict(tweet_vectorized)[0]  # Return prediction as numeric value

def highlight_and_count_discrepancies(df, output_excel):
    wb = Workbook()
    ws = wb.active

    # Append header
    ws.append([col for col in df.columns])  # Assumes df has headers like 'tweet', 'ground_truth', 'classification'

    # Define a red fill style for discrepancies
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    discrepancy_count = 0
    # Write data to the worksheet and highlight discrepancies
    for index, row in df.iterrows():
        ws.append(list(row.values))
        if row['ground_truth'] != row['classification']:
            discrepancy_count += 1
            # Apply fill to last row (index + 2 because of header and zero-based index)
            for cell in ws[index + 2]:
                cell.fill = red_fill

    wb.save(output_excel)
    return discrepancy_count

# Load tweets, add predictions
tweets_df = pd.read_csv('./dataset/testing_dataset/testing-normal.csv')
tweets_df['classification'] = tweets_df['tweet'].apply(predict_tweet)

# Ensure 'ground_truth' exists or simulate it if not present in CSV for demonstration
if 'ground_truth' not in tweets_df.columns:
    tweets_df['ground_truth'] = [0] * len(tweets_df)  # Example placeholder

# Highlight discrepancies and count them
output_path = './test/cross-validation-normal.xlsx'
discrepancy_count = highlight_and_count_discrepancies(tweets_df, output_path)

print(f"Tweet classification complete. Discrepancies highlighted and saved to {output_path}. Total discrepancies: {discrepancy_count}")
